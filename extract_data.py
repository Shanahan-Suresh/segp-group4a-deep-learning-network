import pandas as pd
import numpy as np

pd.options.display.max_columns = None
pd.options.display.max_rows = None


# Opens an excel sheet
def open_sheet(excel):
    import openpyxl

    # saves Excel file as a 'workbook' object, wb
    wb = openpyxl.load_workbook(excel, data_only=True)

    # returns the contents of the first sheet in workbook
    return wb['Sheet1']


# Obtains a table's parameter values from the dataset
def get_parameters(initial_row, initial_column):
    properties = []
    values = []

    for rows in range(initial_row, initial_row + 2):

        for columns in range(initial_column + 1, initial_column + 4):
            cell_value = sheet.cell(row=rows, column=columns).value

            # cell values only saved as a 'property' or 'value' in their respective lists
            if rows == initial_row:
                properties.append(cell_value)

            else:
                values.append([cell_value])

    # create dictionary from completed lists
    parameters = dict(zip(properties, values))

    global df
    df = pd.DataFrame(parameters)  # build dataframe from obtained parameters


# Obtains a table's chemical names, properties, and values from the dataset
def get_chemicals(initial_row, initial_column):
    properties_tag = []
    chemical_names = []
    blank_row = False
    rows = initial_row

    # appends parameter labels into an array
    for columns in range(initial_column, initial_column + 4):

        cell_value = sheet.cell(row=rows, column=columns).value

        if cell_value is None:

            if sheet.cell(row=rows - 2, column=columns - 1).value == "Parameters":
                cell_value = "Composition, g"

            elif sheet.cell(row=rows - 2, column=columns - 2).value == "Parameters":
                cell_value = "SR file number"

            else:
                cell_value = "Chemical Temperature"

        properties_tag.append(cell_value)

    rows += 1

    # appends acid names into an array
    while not blank_row:

        for columns in range(initial_column, initial_column + 4):
            cell_value = sheet.cell(row=rows, column=columns).value
            chemical_names.append([cell_value])

        next_cell_value = sheet.cell(row=rows + 1, column=columns).value

        # exits loop if next cell is empty
        if next_cell_value is None:
            blank_row = True

        rows = rows + 1

        chemicals = dict(zip(properties_tag, chemical_names))  # combine the two arrays into a dictionary structure

        # builds a temporary dataframe from obtained values
        df1 = pd.DataFrame(chemicals)
        bigdata = pd.concat([df, df1], axis=1, join='inner')
        global finaldf
        finaldf = pd.concat([finaldf, bigdata],
                            ignore_index=True)  # concatenates the temporary dataframe into the final dataframe
        del bigdata

        chemical_names.clear()  # clear chemical names for next iteration


# Obtains a table's parameter values from the dataset
def findParameterData():
    global finaldf
    finaldf = pd.DataFrame()

    # hardcoded loop parameters based on dataset
    initial_row = 2
    initial_column = 1
    max_row = sheet.max_row
    max_col = sheet.max_column

    # loops through every row first
    while initial_row != max_row:
        cell_value = sheet.cell(row=initial_row, column=initial_column).value

        # ASSUMES that after finding the word 'parameters', there will be chemicals properties and so on
        if cell_value == "Parameters":
            get_parameters(initial_row, initial_column)  # Obtains parameter values
            get_chemicals(initial_row + 2, initial_column)  # Obtains chemical values

        initial_row += 1

        # ASSUMES that a new table begins every 5 columns
        if initial_row == max_row:
            initial_column += 5

            # exit loop at final column, else row increment by 2
            if initial_column >= max_col:
                break

            else:
                initial_row = 2


# Converts the obtained data into a modified dataframe
def modifyDataset():
    finaldf['Composition, g'].astype(str)
    finaldf['Composition, g'] = finaldf['Composition, g'].fillna(0)
    finaldf.loc[finaldf['Composition, g'].str.contains("Nil", na=False), 'Composition, g'] = 0
    finaldf['Aluminium Temperature'] = finaldf['Aluminium Temperature'].fillna(0)
    finaldf['Chemical Temperature'] = finaldf['Chemical Temperature'].fillna(0)
    finaldf['Chemicals'] = finaldf.Chemicals.astype(str)
    finaldf.loc[
        finaldf['Chemicals'].str.contains("Lauric acid", na=False) & finaldf['Chemicals'].str.contains("Stearic acid",
                                                                                                       na=False), 'Chemicals'] = 4
    finaldf.loc[finaldf['Chemicals'].str.contains("Control", na=False), 'Chemicals'] = 0
    finaldf.loc[finaldf['Chemicals'].str.contains("Lauric acid", na=False), 'Chemicals'] = 1
    finaldf.loc[finaldf['Chemicals'].str.contains("Stearic acid", na=False), 'Chemicals'] = 2
    finaldf.loc[finaldf['Chemicals'].str.contains("Parafin Wax", na=False), 'Chemicals'] = 3

    finaldf['Lauric Acid'] = np.where((finaldf['Chemicals'] == 1) | (finaldf['Chemicals'] == 4), 1, 0)
    finaldf['Stearic Acid'] = np.where((finaldf['Chemicals'] == 2) | (finaldf['Chemicals'] == 4), 1, 0)
    finaldf['Parafin Wax'] = np.where((finaldf['Chemicals'] == 3), 1, 0)
    finaldf['Parafin Wax Composition'] = 0
    finaldf['Lauric Acid Composition'] = 0
    finaldf['Stearic Acid Composition'] = 0

    for i in range(len(finaldf)):

        if (finaldf["Lauric Acid"].iloc[i] == 1) & (finaldf["Stearic Acid"].iloc[i] == 1):

            temparr = finaldf.loc[i, 'Composition, g']
            split = temparr.split("\n")
            finaldf.loc[i, 'Lauric Acid Composition'] = int(split[0])
            finaldf.loc[i, 'Stearic Acid Composition'] = int(split[1])

        elif (finaldf.loc[i, 'Lauric Acid'] == 1) & (finaldf.loc[i, 'Stearic Acid'] != 1):
            finaldf.loc[i, 'Lauric Acid Composition'] = finaldf._get_value(i, 'Composition, g')

        elif (finaldf.loc[i, 'Stearic Acid'] == 1) & (finaldf["Lauric Acid"].iloc[i] != 1):
            finaldf.loc[i, 'Stearic Acid Composition'] = finaldf._get_value(i, 'Composition, g')

        elif finaldf.loc[i, 'Parafin Wax'] == 1:
            finaldf.loc[i, 'Parafin Wax Composition'] = finaldf._get_value(i, 'Composition, g')

    data = finaldf.drop(["Chemicals", "Composition, g"], axis=1)
    shuffle_data = data.sample(frac=1)

    shuffle_data.to_excel("Temp files/Shuffled.xlsx")
    SR_file_numbers = shuffle_data.loc[:, 'SR file number']
    data = shuffle_data.drop(["SR file number"], axis=1)
    return data, SR_file_numbers


# Main Function
def main(excel):
    global sheet  # sheet set as global value
    sheet = open_sheet(excel)
    findParameterData()
    data, SR_file_number = modifyDataset()
    return data, SR_file_number  # returns to neural_network function
